#! /usr/local/bin/python3
""" Create a spreadsheet giving information for each academic program at Queens College.

  Reads a file called 'program_codes.out' generated by a script that scrapes the QC
  information from the NYS DOE website.

  The spreadsheet lines consist of the following columns. The combination of values listed as
  the primary key are guaranteed to be unique.

  Primary Key
    program_code, program_name, HEGIS_code, award, institution
  Fields
    certificate_licenses_titles_types, TAP_eligible, APTS_eligible, WTA_eligible, accreditation

Deductions, based on examination of the program_codes.out file:
  M/A is for Multiple Awards: the same program code may be associated with more than one degree
  M/I is for Multiple Institutions: the same program may be associated with more than one
      institution, which may use different hegis codes and/or awards. If the institution award
      is "NON GRANTING" there is no hegis code for that institution.

Input file structure:
  One program code line
  Zero or more M/I or M/A lines
  For each award:
    For award (BA, MA, etc.)
      Certificate, etc.
      Financial Aid
      Accreditation

Algorithm:
  Read a line, and extract line_type (see Enum line_type)
  Check previous line_type against this line_type to make sure a valid sequence is taking place.
  Then go to the particular algorithm step given below based on this line_type.

  program:                Extract program_code, program_name, hegis_code, award, institution;
                          check for duplicate program code;
  multiple_awards:        Extract program_name, hegis, award, institution. Add to respective sets.
  multiple_instituitions: Extract hegis, award, institution.
  for_award:              Extract award.
  certificate_etc:        Extract certificate tuple {name, type, date} if there is one.
  financial_aid:          Extract three booleans.
  accreditiation:         Extract text, if any.
                          Generate record(s) for this award.
"""
__author__  = 'Christopher Vickery'
__version__ = 'January 2016'

import sys
import re
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('-d', '--debug',    action = 'store_true', default = False)
args = parser.parse_args()

def fix_title(str):
  """ Create a better titlecase string, taking specifics of this dataset into account.
  """
  return str.strip()\
            .title()\
            .replace('Mhc', 'MHC')\
            .replace('\'S', '’s')\
            .replace('1St', '1st')\
            .replace('6Th', '6th')\
            .replace(' And ', ' and ')

""" Input line types, determined from the first token on the line.
"""
from enum import Enum
line_type = Enum('line_type',
                 'program \
                  multiple_awards \
                  multiple_institutions \
                  for_award \
                  certificate_etc \
                  financial_aid \
                  accreditation')

from collections import namedtuple

Key     = namedtuple('Key',
          'program_code program_name hegis_code award institution')
Record  = namedtuple('Record', \
         'cert_name cert_type cert_date tap_eligible apts_eligible vvta_eligible accreditation')
records = dict()

MAW_line  = namedtuple('MAW_line', 'name, hegis, award, institution')
Certs_etc = namedtuple('Certs_etc', 'name type date')
Certs_etc.__new__.__defaults__ = (None,) * len(Certs_etc._fields)

Financial_aid = namedtuple('Financial_aid', 'tap apts vvta')
Financial_aid.__new__.__defaults__ = (False,) * len(Financial_aid._fields)

program_codes = set() # For duplicate checking

this_type     = line_type.accreditation
prev_type     = None
line_num      = 0
num_dupes     = 0

with open('program_codes.out') as fp:
  for line in iter(fp):
    line_num = line_num + 1
    prev_type = this_type
    tokens = line.split()
    token = tokens[0]

    if token.isdecimal():
      # line_type.program
      this_type = line_type.program
      if prev_type is not line_type.accreditation:
        print('line {}: unexpected program code line'.format(line_num))
      # Extract program_code, program_name, hegis_code, award, institution;
      program_code        = token
      program_name        = fix_title(line[10:52])
      program_hegis       = line[52:59]
      program_award       = line[63:76].strip()
      program_institution = line[76:].strip().title().replace('Cuny', 'CUNY')
      # Check for duplicate program code;
      if program_code in program_codes:
        if args.debug:
          print('line {}: duplicate program code: {}'.format(line_num, program_code))
        num_dupes = num_dupes + 1
      program_codes.add(token)

      # Initialize structures for this program code
      maw_lines   = set()
      maw_awards  = set()
      minst_line  = False
      cert        = Certs_etc() # default in case there isn't a certificates line

      if args.debug:
        print('program:',
              program_code, program_name, program_hegis, program_award, program_institution)

    elif token == 'M/A':
      # line_type.multiple_awards
      #
      # There may be multiple M/A (multiple award) lines per program code
      this_type = line_type.multiple_awards
      if prev_type is not line_type.program and \
         prev_type is not line_type.multiple_institutions:
        print('line {}: unexpected M/A line'.format(line_num), file = sys.stderr)
      # Extract name, hegis, award, and institution
      maw_name        = fix_title(line[10:52]).strip()
      maw_hegis       = line[52:59].strip()
      maw_award       = line[63:76].strip()
      maw_institution = line[76:].strip().title().replace('Cuny', 'CUNY')
      maw_line = MAW_line(maw_name, maw_hegis, maw_award, maw_institution)
      if maw_line in maw_lines:
        print('line {}: duplicate M/A ({}) for program code {}'\
              .format(line_num, maw_line, program_code), file = sys.stderr)
      maw_lines.add(maw_line)
      maw_awards.add(maw_award) # For quick lookup
      if args.debug:
        print('M/A:', maw_name, maw_hegis, maw_award, maw_institution)

    elif token == 'M/I':
      # line_type.multiple_institutions
      #
      # Observation: there is never more than one M/I (multiple institutions)
      # line with an award per program code. There might be multiple M/I lines
      # but no more than one of them has an award. If this ever changes, this
      # section of the code will have to be updated to work like M/A lines.
      this_type = line_type.multiple_institutions
      if prev_type is not line_type.program and \
         prev_type is not line_type.multiple_awards:
        print('line {}: unexpected M/I line'.format(line_num), file = sys.stderr)
      # Extract hegis, award, institution
      minst_hegis = line[52:59].strip()
      if minst_hegis == "":
        minst_hegis = None
      minst_award = line[63:76].strip()
      if minst_award == 'NOT-GRANTING':
        minst_award = None
      if minst_award:
        if minst_line:
          print('line {}: second M/I line for program code {}'.format(line_num, program_code))
          exit() # fatal
        minst_institution = line[76:].strip().title().replace('Cuny', 'CUNY')
        minst_line = True
        if args.debug:
          print('M/I:', minst_hegis, minst_award, minst_institution)

    elif token == 'FOR':
      # line_type.for_award
      #
      this_type = line_type.for_award
      if prev_type is not line_type.program and \
         prev_type != line_type.multiple_awards and \
         prev_type != line_type.multiple_institutions and\
         prev_type != line_type.accreditation:
        print('line {}: unexpected for award line'.format(line_num), file = sys.stderr)
      # Extract award (is it in the set of awards extracted from program and m/a lines?)
      for_award = re.match('\s*FOR AWARD\s*--(.*)', line).group(1).strip()
      if for_award != program_award and for_award not in maw_awards:
        # fatal
        print('line {}: for award ({}) not in awards ({})'\
              .format(line_num, for_award, awards), file = sys.stderr)
        exit()

    elif token.startswith('CERT'):
      # line_type.certificate_etc
      #
      this_type = line_type.certificate_etc
      if prev_type != line_type.for_award:
        print('line{}: unexpected certificate_etc line'.format(line_num), file = sys.stderr)
      # Extract certificate tuple {name, type, date} if there is one.
      cert_info = line[47:].strip()
      if cert_info.startswith('NONE'):
        cert = Certs_etc()
      else:
        cert = Certs_etc(cert_info[0:18].strip(), cert_info[18:28].strip(), cert_info[28:].strip())

    elif token == 'PROGRAM' and tokens[1] == 'FINANCIAL':
      # line_type.financial_aid
      #
      this_type = line_type.financial_aid
      if prev_type != line_type.certificate_etc and\
         prev_type != line_type.for_award:
        print('line {}: unexpected financial_aid line'.format(line_num), file = sys.stderr)
      # Extract three booleans.
      if_tap  = line[54:57] == 'YES'
      if_apts = line[66:69] == 'YES'
      if_vvta = line[77:80] == 'YES'
      fin_aid = Financial_aid(if_tap, if_apts, if_vvta)

    elif token == 'PROGRAM' and tokens[1] == 'PROFESSIONAL':
      # line_type.accreditation
      #
      this_type = line_type.accreditation
      if prev_type != line_type.financial_aid:
        print('line {}: unexpected accreditation line'.format(line_num), file = sys.stderr)
      # Extract text, if any.
      accreditation = line[45:].strip()

      # Observation: every program's section ends with an accreditation line.
      #
      # Generate record(s) for this program.
      #

      # Assemble the data record
      #   cert_name cert_type cert_date tap_eligible apts_eligible vvta_eligible accreditation
      data_record = Record(cert.name,   cert.type,    cert.date,
                           fin_aid.tap, fin_aid.apts, fin_aid.vvta,
                           accreditation)

      # Generate all distinct keys for this award for this program code
      keys = set()

      # Multiple award sets, if any
      if for_award in maw_awards:
        for maw_line in maw_lines:
          if maw_line.award == for_award:
            key = Key(program_code,
                      maw_line.name,
                      maw_line.hegis,
                      maw_line.award,
                      maw_line.institution)
            if key in records.keys():
              if args.debug:
                print('Duplicate key while procssing M/A for program code {}:\n  {}'\
                      .format(program_code, key))
              if records[key] != data_record:
                print('Duplicated maw key with different data: {}\n  {}\n  {}'.format(
                    key, records[key], data_record), file = sys.stderr)
            else:
              records[key] = data_record

      # Multiple institutions, if any
      if minst_line and minst_award == for_award:
        key = Key(program_code,
                  re.sub('  +', ' ', program_name),
                  minst_hegis,
                  minst_award,
                  minst_institution)
        if key in records.keys():
          if args.debug:
            print('Duplicate key while processing M/I for program code {}:\n  {}'\
                  .format(program_code, key))
          if records[key] != data_record:
            print('Duplicated M/I key with different data: {}\n  {}\n  {}'.format(
                key, records[key], data_record), file = sys.stderr)
        else:
          records[key] = data_record

      # Programs without multiple institutions or awards
      if program_award == for_award:
        key = Key(program_code,
                  re.sub('  +', ' ', program_name),
                  program_hegis,
                  program_award,
                  program_institution)
        if key in records.keys():
          if args.debug:
            print('Duplicate key while processing program code {}:\n  {}'\
                  .format(program_code, key))
          if records[key] != data_record:
            print('Duplicated program key with different data: {}\n  {}\n  {}'.format(
                key, records[key], data_record))
        else:
          records[key] = data_record
    else:
      print('line {}: unrecognized line type:\n  -->{}<--'\
            .format(line_num, line), file = sys.stderr)

#
# Generate spreadsheet
#
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active
ws.title = 'Academic Programs'
row = 1
headers = 'Program Code', 'Program Name', 'HEGIS', 'Award', 'Institution', \
            'Certificate Name', 'Certificate Type', 'Certificate Date', \
            'TAP Eligible', 'APTS Eligible', 'VVTA Eligible'
bold = Font(bold = True)
for col in range(len(headers)):
  cell = ws.cell(row = row, column = col + 1)
  cell.value = headers[col]
  cell.font = bold
for key in sorted(records.keys()):
  row = row + 1
  for col in range(len(key)):
    cell = ws.cell(row = row, column = col + 1)
    cell.value = key[col]
  for col in range(len(records[key])):
    cell = ws.cell(row = row, column = col + len(key) + 1)
    cell.value = records[key][col]
wb.save('program_codes.xlsx')
print('Processed {} lines.\nFound {} distinct program codes and {} duplicate program codes.'\
      .format(line_num, len(program_codes), num_dupes))
print('Generated', len(records), 'records.')
